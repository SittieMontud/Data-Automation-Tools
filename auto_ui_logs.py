import openpyxl
import re
import helpers
import time
import tkinter


def parsing(stringPathEntry, outputtextUI):
	#Configure time
	timestamp = time.ctime()
	timestamp = timestamp.replace("Wed","")
	timestamp = timestamp.replace(" ","_")
	timestamp = timestamp.replace(":","_")
	
	#Create workbook
	filepath = stringPathEntry
	workbookPath = filepath.replace(".txt","_") + "Parsed_" + timestamp + ".xlsx"
	openpyxl.Workbook().save(workbookPath)
	workbookAct = openpyxl.load_workbook(workbookPath)
	
	#Input number of controllers
	#maxCont = int(stringContEntry)
	
	#Creating and changing the name of the sheets
	sheetlist = workbookAct.sheetnames
	sheetname = sheetlist[0]
	sheet = workbookAct[sheetname]
	sheet.title = "Parsed Output"
	workbookAct.create_sheet("Test Limits")
	sheet2 = workbookAct['Test Limits']
	
	stopTestname = 0

	#Open the text file
	textpath = stringPathEntry
	outputtextUI.insert(tkinter.END, "-Opening file %s " % textpath)
	datalog = open(textpath, "r")

	outputtextUI.insert(tkinter.END, "\n-Parsing..." )
	
	#Store all data in the text file to dataList
	dataList = []
	datalog.seek(0)
	for data in datalog.readlines():
		dataList.append(data)
	maxLine = len(dataList)

	#Get Start and End Flags
	dataSample = dataList
	stringAll = ""
	for rows in range(maxLine):
		dataSample[rows].rstrip('\n').split(' ')
		for elem in dataSample[rows]:
			stringAll += elem

	flagStart = re.findall(r" Started Test Cycle Count <\d+> ...\n", stringAll)
	flagEnd = re.findall(r" Completed Test Cycle Count <\d+> ...\n", stringAll)
	#Checker for start and end flags
	if (len(flagStart) == len(flagEnd)):
		print("%d loop(s) found in the file" % len(flagStart))
	else:
		print("Mismatched flagStart and flagEnd.")
	
	#Check number of DUTs
	dutList = re.findall(r"DUT: \d   \w", stringAll)
	dutString = helpers.ListToString(dutList)
	dutList = re.findall(r"\d", dutString)
	dutList = sorted(helpers.RemoveDuplicate(dutList))
	
	
	#Iterate per cycle count
	startRange = 0
	endRange = 0
	helpers.startloop = 1
	for loop in range(len(flagStart)):
		print("--Parsing loop %d" % helpers.startloop)
		
		#Get tests per row and store in dataTests
		dataTests = []
		startRange = dataList.index(flagStart[loop]) + 1
		endRange = dataList.index(flagEnd[loop]) + 1
		for data in range(startRange, endRange):
			dataTests.append(dataList[data])
		
		#Create LIST of the dut informations
		dieLoc = ""
		for test in dataTests:
			dieLoc += test	
		dict_dut = {}
		for num in range(0,len(dutList)):
			dutloc = re.findall(r"DUT: \d.+", dieLoc)[num]
			dict_dut[int(dutList[num])] = dutloc
			dict_dut[int(dutList[num])] = helpers.ListToStringToList(dutloc)
			emptyList = []
			for elem in dict_dut[int(dutList[num])]:
				if elem != '':
					emptyList.append(elem)
			dict_dut[int(dutList[num])] = emptyList
		
		#Get Test Names HEADERS and Test limits then copy to excel
		if stopTestname == 0:
			testNames = []
			testLSL = {}
			testUSL = {}
			#Loop for testnames
			for test in dataTests:
				strRow = ""
				strRow = helpers.RemoveSpaces(test, strRow)
				if len(strRow.split(' ')) > 14:
					dataRow = helpers.UIMakeTestName(strRow)
					testNames.append(dataRow[3])					
					testLSL[dataRow[3]] = []
					testUSL[dataRow[3]] = []
			#Loop for testlimits
			for test in dataTests:
				strRow = ""
				strRow = helpers.RemoveSpaces(test, strRow)
				if len(strRow.split(' ')) > 14:
					dataRow = helpers.UIMakeTestName(strRow)
					limitlist = testLSL[dataRow[3]]
					limitsL = float("{:.3f}".format(float(helpers.UILowTestLimits(strRow))))
					limitlist.append(limitsL)
					limitlist = testUSL[dataRow[3]]
					limitsU = float("{:.3f}".format(float(helpers.UIHighTestLimits(strRow))))
					limitlist.append(limitsU)
	
			# Removing duplicate test names
			testNames = helpers.RemoveDuplicate(testNames)
			#Making list of test limits
			testnamesLimit = []
			for name in testNames:
				testnamesLimit.append(name)
			testnamesLimit.insert(0, "TEST NAMES")
			listLSL = []
			listUSL = []
			for keys in testLSL:
				listLSL.append(testLSL[keys][0])
			for keys in testUSL:
				listUSL.append(testUSL[keys][0])
			listLSL.insert(0, "LOWER LIMIT")
			listUSL.insert(0, "UPPER LIMIT")
	
			# Copying test limits to excel
			sheet2.append(testnamesLimit)
			sheet2.append(listUSL)
			sheet2.append(listLSL)
			#Copying test names to excel
			testNames.insert(0, 'TOUCHDOWN')
			testNames.insert(1, 'CONTROLLER')
			testNames.insert(2, 'DUT')
			testNames.insert(3, 'DIE X')
			testNames.insert(4, 'DIE Y')
			testNames.insert(5, 'SOFT BIN')
			testNames.insert(6, 'HARD BIN')
			sheet.append(testNames)
	
		#Get Values and inserting dut number to list
		testResults = []
		listVal = {}
		for num in range(0,len(dutList)):
			listVal[dutList[num]] = [int(dutList[num])] #Inserting DUT number
			listVal[dutList[num]].insert(0, helpers.startloop)	#Inserting touchdo wn to list
			listVal[dutList[num]].insert(1, 1)	#Inserting controller to list
			listVal[dutList[num]].insert(3, int(dict_dut[int(dutList[num])][3])) 	#Inserting DIE X
			listVal[dutList[num]].insert(4, int(dict_dut[int(dutList[num])][5])) 	#Inserting DIE Y
			listVal[dutList[num]].insert(5, int(dict_dut[int(dutList[num])][7])) 	#Inserting SOFT BIN
			listVal[dutList[num]].insert(6, int(dict_dut[int(dutList[num])][9])) 	#Inserting HARD BIN
			for test in dataTests:
				strRow = ""
				strRow = helpers.RemoveSpaces(test, strRow)
				if len(strRow.split(' ')) > 14: #Functional Tests not included
					listdummy = strRow.split(' ')
					if (int(dutList[num]) == int(listdummy[1])):
						measvalues = float("{:.3f}".format(float(helpers.UIStoreValues(strRow))))
						listVal[dutList[num]].append(measvalues)
			sheet.append(listVal[dutList[num]])	#Copy list to excel
		helpers.startloop += 1
		stopTestname += 1
	

	#Save workbook
	#outputtextASCII.insert(tkinter.END,"\n-->Saving excel file to %s" % workbookPath)
	outputtextUI.insert(tkinter.END,"\n-Saving excel file")
	workbookAct.save(workbookPath)
	outputtextUI.insert(tkinter.END,"\n-Parsing Done!\n")

