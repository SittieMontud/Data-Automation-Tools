import openpyxl
from openpyxl.utils import get_column_letter
import re
import time
import math
import tkinter

def converttoexcel(pathsStr, outputtextTT):
	#Configure time
	timestamp = time.ctime()
	timestamp = timestamp.replace("Wed","")
	timestamp = timestamp.replace(" ","_")
	timestamp = timestamp.replace(":","_")
	
	#Convert String to a list
	pathList = pathsStr.split("\n")
	pathcontList = []
	for path in pathList:
		if path != "":
			pathcontList.append(path)
	pathList = pathcontList
	
	#Create workbook
	pathdir = re.search(r"\w+.txt", pathList[0]).group(0)
	workbookPath = pathList[0].strip(pathdir) + "Test Time_" + timestamp + ".xlsx"
	openpyxl.Workbook().save(workbookPath)
	workbookAct = openpyxl.load_workbook(workbookPath)
	
	ipath = 0
	for pathfile in pathList:
		#Creating and changing the name of the sheets
		textname = re.search(r"\w+.txt", pathfile).group(0)
		if ipath == 0:
			sheetlist = workbookAct.sheetnames
			sheetname = sheetlist[0]
			sheet = workbookAct[sheetname]
			sheet.title = textname.rstrip(".txt")
		else:
			workbookAct.create_sheet(textname.rstrip(".txt"))
			sheet = workbookAct[textname.rstrip(".txt")]
		
		outputtextTT.insert(tkinter.END, "-->Opening file %s" % pathfile)
		#Open test file
		ttlog = open(pathfile, "r")
		
		#Store contents to a list
		datalist = []
		ttlog.seek(0)
		for data in ttlog.readlines():
			datalist.append(data)
		
		#Convert content list to string
		datastr = ""
		for data in datalist:
			datastr += data
		
		#Find start and end flags
		startmatch = re.findall(" Started Test Cycle Count <.+> ...\n", datastr)
		endmatch = re.findall(" Completed Test Cycle Count <.+> ...\n", datastr)
		
		#Length of start flags
		startlen = len(startmatch)
		
		datacyclelist = []
		datacycledummylist = []
		datacyclestr = ""
		containerlist = []
		containerlist2 = []
		containerlist3 = []
		containerlist4 = []
		containterstr = ""
		testtimedict = {}
		ix = 0
		for loop in range(startlen): #startlen
			startindex = datalist.index(startmatch[loop])
			endindex = datalist.index(endmatch[loop])
			for data in datalist[startindex: endindex+1]:
				containerlist.append(data)
			for data in containerlist:
				containerlist2 = data.strip('\n').split(" ")
				for elem in containerlist2:
					if elem != "":
						containerlist3.append(elem)
				datacyclelist.append(containerlist3)
				#Clear lists
				containerlist2 = []
				containerlist3 = []
			
			#Remove empty list
			for data in datacyclelist:
				if data != []:
					datacycledummylist.append(data)
			datacyclelist = datacycledummylist
		
			datacyclelen = len(datacyclelist)
			for num in range(datacyclelen):
				if ix == 0:
					if datacyclelist[num][0] == "Test":
						testtimedict[datacyclelist[num][2]] = [datacyclelist[num][3]]
				else:
					if datacyclelist[num][0] == "Test":
						containerlist4 = testtimedict[datacyclelist[num][2]]
						containerlist4.append(datacyclelist[num][3])
						testtimedict[datacyclelist[num][2]] = containerlist4
			#Clear list
			datacyclelist = []
			containerlist = []
			datacycledummylist = []
			
			#Increment loop
			ix += 1
		
		headerlist = []
		testtimekeys = testtimedict.keys()
		testtimekeyslist = list(testtimekeys)
		headernum = len(testtimedict[testtimekeyslist[0]])
		
		#Make header and copy to excel
		headerlist = ["TEST NAME"]
		for num in range(1,headernum+1):
			headerlist.append(num)
		headerlist.append("AVERAGE")
		sheet.append(headerlist)
		
		#Copy data to excel
		outputlist = []
		for keys in testtimekeyslist:
			outputlist =  [keys]
			for values in testtimedict[keys]:
				outputlist.append(float(values))
			sheet.append(outputlist)
			outputlist = []
		
		#Sum and Average
		averageloop = headernum
		sumlist = []
		avelist = []
		#Sum
		maxcol = sheet.max_column
		maxrow = sheet.max_row
		for col in range(2, maxcol):
			sum = 0.00
			colletter = get_column_letter(col)
			for row in range(2, maxrow + 1):
				sum += float(sheet["%s%s" % (colletter, row)].value)
			sum = float("{:.3f}".format(sum))
			sumlist.append(sum)
		sumlist.insert(0, "TOTAL TEST TIME")
		sheet.append(sumlist)
		#Average
		maxcol = sheet.max_column
		maxrow = sheet.max_row
		avecol = get_column_letter(maxcol) 
		for row in range(2, maxrow + 1):
			ave = 0.0
			sum = 0.0
			for col in range(2, maxcol):
				colletter = get_column_letter(col)
				sum += float(sheet["%s%s" % (colletter, row)].value)
			ave = sum / averageloop
			ave = float("{:.3f}".format(ave))
			sheet["%s%s" % (avecol, row)].value = ave
		
		#Change font
		maxcol = sheet.max_column
		maxrow = sheet.max_row
		fontvar = openpyxl.styles.Font(bold = True, italic = False)
		fontvarValue = openpyxl.styles.Font(color = '00FF0000', bold = True, italic = False)
		for col in range(1, maxcol + 1):
			colletter = get_column_letter(col)
			sheet["%s1" % colletter].font = fontvar
		for row in range(1, maxrow + 1):
			sheet["A%s" % row].font = fontvar
		for row in range(2, maxrow + 1):
			sheet["%s%s" % (get_column_letter(maxcol), row)].font = fontvarValue
		
		#Increment sheet
		ipath += 1
	
	#Save workbook
	outputtextTT.insert(tkinter.END, "\n-->Saving file to %s" % workbookPath)
	workbookAct.save(workbookPath)
	
	outputtextTT.insert(tkinter.END, "\n-->Conversion to excel file done!")


















